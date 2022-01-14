import numpy as np
import sys
#control is 's' or 'e', value is magnitude
import openpyxl as xl
from Classes.exp import experiment as exp


def exp_ld(time_factor = 1.0):
	wb = xl.load_workbook('2V9.xlsx',data_only = True)
	sh = wb['11JUN']
	time = 0.
	dt = sh.cell(row = 8, column = 14).value - sh.cell(row = 7, column = 14).value
	#dt = np.array([])
	#time_exp = exp.get_time()
	#for k in range(len(time_exp)-1):
	#	dt = np.append(dt, time_exp[k+1]-time_exp[k])
	#print(np.min(dt),np.max(dt))
	#sys.exit()
	#
	stretch = 0.0
	F = np.eye(3)
	GL = np.zeros((3,3))
	GL_11 = np.array([])
	delta_GL11 = np.array([0.0])
	#
	
	for i in range(7,11907,1):
		#dtime = sh.cell(row = i, column = 14).value - sh.cell(row = i-1, column = 14).value
		#dt = np.append(dt, dtime)
		#print(np.max(dt),np.min(dt))
		#sys.exit()
	#
		stretch = sh.cell(row = i,column = 13).value 
		#
		F = np.array([[stretch,0.,0.],[0.,1./np.sqrt(stretch),0.],[0.,0.,1./np.sqrt(stretch)]])
		Fi = np.linalg.inv(F) # inverse of deformation gradient
		#
		GL = 0.5*(np.dot(F.T,F) - np.identity(3))
		GL_11 = np.append(GL_11, GL[0,0])
	for j in range(len(GL_11)-1):
		delta_GL11 = np.append(delta_GL11, GL_11[j+1]-GL_11[j])
	#
	loading_scheme = ['e','e','e','e','e','e']
	load = np.array([])
	#
	load = [[ time_factor*dt, time_factor*dt,[g,0.,0.,0.,0.,0.], loading_scheme ]  for g in delta_GL11 ]
	#
	
	print("called 1 time")
	#
	return load	

#############################################################################################################

def experiment_ld(time_factor = 1.0):
	dt = np.array([0.0])
	time_exp = exp.get_time()
	for i in range(len(time_exp)-1):
		dt = np.append(dt, time_exp[i+1]-time_exp[i])
	delta_GL11 = np.array([0.0])
	strain_exp_11 = exp.get_strain_11()
	for j in range(len(strain_exp_11)-1):
		delta_GL11 = np.append(delta_GL11, strain_exp_11.T[j+1]-strain_exp_11.T[j])
	loading_scheme = ['e','s','s','e','e','e']
	load_step = [0.0,0.0,np.array([0.0]*6),loading_scheme]
	#
	load = [[time_factor*dt[k], time_factor*dt[k],[delta_GL11[k],0.,0.,0.,0.,0.],loading_scheme] for k in range(len(dt))]
	#print("called 1 time")
	return load

#############################################################################################################


def experiment_creep(time_factor = 1.0):
	dt = np.array([0.0])
	time_exp = exp.get_time()
	for i in range(len(time_exp)-1):
		dt = np.append(dt, time_exp[i+1]-time_exp[i])
	#delta_GL11 = np.array([0.0])
	#strain_exp_11 = exp.get_strain_11()
	#for j in range(len(strain_exp_11)-1):
	#	delta_GL11 = np.append(delta_GL11, strain_exp_11.T[j+1]-strain_exp_11.T[j])
	loading_scheme = ['s','s','s','s','s','s']
	load_step = [0.0,0.0,np.array([0.0]*6),loading_scheme]
	#
	load_init = [[0.0,0.0,[0.,0.,0.,0.,0.,0.],['s','s','s','s','s','s']],[1.0,1.0,[-7.2,0.,0.,0.,0.,0.],['s','s','s','s','s','s']]]
	load_unload = [[0.0,0.0,[0.,0.,0.,0.,0.,0.],['s','s','s','s','s','s']],[1.0,1.0,[7.2,0.,0.,0.,0.,0.],['s','s','s','s','s','s'],[1.0,1.0,[7.2,0.,0.,0.,0.,0.],['s','s','s','s','s','s']]]
	#
	#load = [[time_factor*dt[k], time_factor*dt[k],[0.,0.,0.,0.,0.,0.],loading_scheme] for k in range(len(dt))]
	#load = load_init+load
	load = load_init+load_unload

	#print("called 1 time")
	return load

#############################################################################################################



# Experiment load case from exp.py

def experiment_load(time_factor = 1.0):
	dt = np.array([0.0])
	time_exp = exp.get_time()
	for i in range(len(time_exp)-1):
		dt = np.append(dt, time_exp[i+1]-time_exp[i])
	delta_GL11 = np.array([0.0])
	delta_GL22 = np.array([0.0])
	delta_GL33 = np.array([0.0])
	strain_exp_11 = exp.get_strain_11()
	strain_exp_22 = exp.get_strain_22()
	strain_exp_33 = exp.get_strain_33()
	for j in range(len(strain_exp_11)-1):
		delta_GL11 = np.append(delta_GL11, strain_exp_11.T[j+1]-strain_exp_11.T[j])
		delta_GL22 = np.append(delta_GL22, strain_exp_22.T[j+1]-strain_exp_22.T[j])
		delta_GL33 = np.append(delta_GL33, strain_exp_33.T[j+1]-strain_exp_33.T[j])
	loading_scheme = ['e','e','e','e','e','e']
	load_step = [0.0,0.0,np.array([0.0]*6),loading_scheme]
	#
	#load = [[time_factor*dt[k], time_factor*dt[k], strain_exp[k],loading_scheme] for k in range(len(dt))]
	load = [[time_factor*dt[k], time_factor*dt[k],[delta_GL11[k],delta_GL22[k],delta_GL33[k],0.,0.,0.],loading_scheme] for k in range(len(dt))]
	print("called 1 time")
	return load

#[delta_GL11[k],0.,0.,0.,0.,0.]
#############################################################################################################

def relax(time_factor = 1.): # ramp with single relaxation
	# t
	t = np.array([])
	t = np.linspace(0,10,1010)
	dt = (max(t)-min(t))/len(t)
	# lambda
	total_stretch = 0.95
	
	GL_inp = delta_GL11(t,total_stretch)
	#
	# ramp
	#
	#print(GL_inp)
	loading_scheme = ['e','e','e','e','e','e']
	load = np.array([])
	#
	load = [ [ time_factor*dt, time_factor*dt, [g, 0.,0.,0.,0.,0.], loading_scheme ]  for g in GL_inp ]
	#
	#print(len(load))
	return load



def delta_GL11(t,total_stretch): # for single ramp and relaxation
	dt = (np.max(t)-np.min(t))/len(t)
	dlambda = (total_stretch - 1.)/len(t)
	stretch = 1. - 2.*dlambda
	delta_GL11 = np.array([0.])
	GL_11 = np.array([])
	for i in range(len(t)):
		if (stretch > total_stretch):
			stretch += 2.*dlambda
		CG = np.diag([stretch**2.,1.,1.])
		GL = 0.5*(CG - np.identity(3))
		GL_11 = np.append(GL_11, GL[0,0])
	for j in range(len(GL_11)-1):
		delta_GL11 = np.append(delta_GL11,GL_11[j+1]-GL_11[j])
	return delta_GL11
####################################################################

def relax_steps(time_factor = 1.,total_stretch = 0.95,t_max=30., dt = 0.1,dlambda = -0.6): # ramp with single relaxation
	# t
	#t = np.array([])
	#t = np.linspace(0,10,1010)
	#dt = (max(t)-min(t))/len(t)
	#t_max = 10.
	nsteps = np.int(t_max/dt)
	t = np.linspace(0,t_max, nsteps)
	# lambda
	#total_stretch = 0.95
	
	GL_inp = delta_GL11_steps(dt,t_max,total_stretch,dlambda) # dlambda is the stretch rate
	#
	# ramp
	#
	#print(GL_inp)
	loading_scheme = ['e','e','e','e','e','e']
	load = np.array([])
	#
	load = [ [ time_factor*dt, time_factor*dt, [g, 0.,0.,0.,0.,0.], loading_scheme ]  for g in GL_inp ]
	#
	#print(len(load))
	return load



def delta_GL11_steps(dt,t_max,total_stretch,dlambdadt): # for single ramp and relaxation 
	#dt = (np.max(t)-np.min(t))/len(t)
	#dlambda = (total_stretch - 1.)/t_max
	nsteps=np.int(t_max/dt)
	stretch = 1. 
	delta_GL11 = np.array([0.])
	GL_11 = np.array([])
	counter = 1
	for i in range(nsteps):
		if (stretch > total_stretch):
			stretch = 1.+ dlambdadt*dt*i # dlambdadt is the stretch rate
			counter+=1
		CG = np.diag([stretch**2.,1.,1.])
		GL = 0.5*(CG - np.identity(3))
		GL_11 = np.append(GL_11, GL[0,0])
	print(counter)
	for j in range(len(GL_11)-1):
		delta_GL11 = np.append(delta_GL11,GL_11[j+1]-GL_11[j])
	return delta_GL11	
	
#######################################################################################
	
def delta_GL_11(t,dlambda,stretch,total_stretch): # asisting function for ramp_relax
	dt = (np.max(t)-np.min(t))/len(t)
	#dlambda = (total_stretch - 1.)/len(t)
	#stretch = 1.-4.*dlambda
	delta_GL11 = np.array([0.])
	GL_11 = np.array([])
	for i in range(len(t)):
		if (stretch > total_stretch):
			stretch += 4.*dlambda
		CG = np.diag([stretch**2.,1.,1.])
		GL = 0.5*( CG - np.identity(3))
		#sys.exit()
		GL_11 = np.append(GL_11, GL[0,0])
		#delta_GL11 = np.append(delta_GL11, GL[0,0])
	#print('load case',GL_11, len(GL_11))
	#print(GL_11[1009],GL_11[0],len(delta_GL11))
	#
	for j in range(len(GL_11)-1):
		delta_GL11 = np.append(delta_GL11,GL_11[j+1]-GL_11[j])
	#print(delta_GL11,len(delta_GL11))
	#sys.exit()
	#
	return delta_GL11
	
	
def GL11_ld(t,dlambda,stretch,total_stretch): # asisting function for ramp_uld
	dt = (np.max(t)-np.min(t))/len(t)
	GL11_ld = np.array([])
	GL_11 = np.array([])
	for i in range(len(t)):
		if (stretch>total_stretch):
			stretch += 1.*dlambda #change here for relaxation or not, with relaxation = 4.*, without = 1.*
			CG = np.diag([stretch**2.,1.,1.])
			GL = 0.5*( CG - np.identity(3))
			#sys.exit()
			GL_11 = np.append(GL_11, GL[0,0])
	for j in range(len(GL_11)-1):
		GL11_ld = np.append(GL11_ld,GL_11[j+1]-GL_11[j])
	#print(delta_GL11,len(delta_GL11))
	#sys.exit()
	#
	return GL11_ld


def ramp_relax(time_factor = 1.): #ramping with nramp relaxations + unloading with nunld relaxations: change nramp and nunld values
	nramp = 1
	# t
	#t = np.array([])
	t = np.linspace(0,10,1010)
	dt = (max(t)-min(t))/len(t)  # change to max(t)-min(t)
	# lambda
	total_stretch = 0.92
	sramp = (total_stretch - 1.)/nramp # step ramp
	#
	ramp_stretch = np.array([0.]*nramp)
	GL_inp = np.array([])
	dlambda = 0.
	stretch = 1.
	for k in range(nramp):
		ramp_stretch[k] = 1.+sramp*(k+1)
		#print(ramp_stretch[k],k,t,len(t))
		dlambda = np.where(k<1, (ramp_stretch[k]-1.)/len(t), (ramp_stretch[k]-ramp_stretch[k-1])/len(t)) # uncomment when ramping with relaxations
		stretch = np.where (k<1,(1.-4.*dlambda), (ramp_stretch[k-1]-4.*dlambda)) #uncomment when ramping with relaxations
		GL_inp = np.append(GL_inp,delta_GL_11(t,dlambda, stretch, ramp_stretch[k])) # uncomment when ramping with relaxations
		#print(GL_inp,len(GL_inp))
		#t += 1.*np.max(t)/(k+1) # no need for t, as dt is passed with delta_E11
		#print(dlambda,ramp_stretch[k])
		#
	#print(t,len(t))
	#sys.exit()
	################### unloading ################################
	#
	#     one step unloading
	#
	############################################
	#dlambda_uld = (1.-total_stretch)/(len(t)*nramp) #uld = unloading
	#stretch_uld = total_stretch + 4.*dlambda_uld
	#
	#GL_inp = np.append(GL_inp,delta_GL_11(100*t,dlambda_uld,stretch_uld,total_stretch)) # single step unloading
	#
	###########################################
	#
	#    nuld step unloading 
	#
	############################################
	nuld = 4 # number of unloading-relaxation steps
	uramp = (1. -total_stretch)/nuld 
	unld_stretch = np.array([0.]*nuld)
	GL11_uld = np.array([])
	delta_GL11_uld = np.array([])
	for i in range(nuld):# nuld = number of unload relaxations
		unld_stretch[i] = total_stretch + uramp*(i+1)
		uld_dlambda = np.where(i<1, (unld_stretch[i]-total_stretch)/len(t), (unld_stretch[i]-unld_stretch[i-1])/len(t)) 
		uld_stretch = np.where(i<1, (total_stretch - 4.*uld_dlambda),(unld_stretch[i-1]- 4.*uld_dlambda))
		print(unld_stretch[i],i,total_stretch)
		#print(uld_stretch,uld_dlambda)
		#
		#if (i<1):
		#	print(np.greater(uld_stretch, total_stretch))
		#
		for j in range(len(t)):
			if (uld_stretch < unld_stretch[i]):
				uld_stretch += 4. *uld_dlambda
			CG_uld = np.diag([uld_stretch**2.,1.,1.])
			GL_uld = 0.5 * (CG_uld - np.identity(3))
			GL11_uld = np.append(GL11_uld,GL_uld[0,0])
		#print(len(t),len(GL11_uld),len(delta_GL11_uld)) # checked len(GL11_uld) correct
	for m in range(len(GL11_uld)-1):
		delta_GL11_uld = np.append(delta_GL11_uld, GL11_uld[m+1]-GL11_uld[m])
	#print(len(delta_GL11_uld))
	#sys.exit()
	#
	GL_inp = np.append(GL_inp,delta_GL11_uld)
	
	############################################
	#
	#sys.exit()		
	#print(GL_inp[0],GL_inp[1010],GL_inp[2020],GL_inp[3030])#,GL_inp[4040],GL_inp[5050])
	loading_scheme = ['e','e','e','e','e','e']
	load = np.array([])
	#
	#t = np.append(t,np.append(t*2, np.append(t*3,t*4)))
	load = [ [ dt, dt, [g, 0.,0.,0.,0.,0.], loading_scheme ]  for g in GL_inp ]
	#
	print(len(load))
	return load


def ramp_uld(time_factor = 1.): # ramping_unloading_relax
	nramp = 1
	# t
	#t = np.array([])
	t = np.linspace(0,10,101)
	dt = (max(t)-min(t))/len(t)
	# lambda
	total_stretch = 0.95
	sramp = (total_stretch - 1.)/nramp # step ramp
	#
	ramp_stretch = np.array([0.]*nramp)
	GL_inp = np.array([])
	dlambda = 0.
	stretch = 1.
	for k in range(nramp):
		#
		##########################################################################
		# ramping without relaxation
		#
		dlambda = (total_stretch - 1.)/len(t)
		stretch = 1.-1.*dlambda
		GL_inp = np.append(GL_inp,GL11_ld(t,dlambda, stretch, ramp_stretch[k]))
		#
	#print(t,len(t))
	#sys.exit()
	################### unloading ################################
	#
	#     one step unloading
	#
	############################################
	#dlambda_uld = (1.-total_stretch)/(len(t)*nramp) #uld = unloading
	#stretch_uld = total_stretch + 4.*dlambda_uld
	#
	#GL_inp = np.append(GL_inp,delta_GL_11(100*t,dlambda_uld,stretch_uld,total_stretch)) # single step unloading
	#
	###########################################
	#
	#    nuld step unloading 
	#
	############################################
	nuld = 1 # number of unloading-relaxation steps
	uramp = (1. -total_stretch)/nuld 
	unld_stretch = np.array([0.]*nuld)
	GL11_uld = np.array([])
	delta_GL11_uld = np.array([])
	for i in range(nuld):# nuld = number of unload relaxations
		unld_stretch[i] = total_stretch + uramp*(i+1)
		uld_dlambda = np.where(i<1, (unld_stretch[i]-total_stretch)/len(t), (unld_stretch[i]-unld_stretch[i-1])/len(t)) 
		uld_stretch = np.where(i<1, (total_stretch - 4.*uld_dlambda),(unld_stretch[i-1]- 4.*uld_dlambda))
		print(unld_stretch[i],i,total_stretch)
		#print(uld_stretch,uld_dlambda)
		#
		#if (i<1):
		#	print(np.greater(uld_stretch, total_stretch))
		#
		for j in range(len(t)):
			if (uld_stretch < unld_stretch[i]):
				uld_stretch += 4. *uld_dlambda
			CG_uld = np.diag([uld_stretch**2.,1.,1.])
			GL_uld = 0.5 * (CG_uld - np.identity(3))
			GL11_uld = np.append(GL11_uld,GL_uld[0,0])
		#print(len(t),len(GL11_uld),len(delta_GL11_uld)) # checked len(GL11_uld) correct
	for m in range(len(GL11_uld)-1):
		delta_GL11_uld = np.append(delta_GL11_uld, GL11_uld[m+1]-GL11_uld[m])
	print(len(delta_GL11_uld))
	#sys.exit()
	#
	GL_inp = np.append(GL_inp,delta_GL11_uld)
	
	############################################
	#
	#sys.exit()		
	#print(GL_inp[0],GL_inp[1010],GL_inp[2020],GL_inp[3030])#,GL_inp[4040],GL_inp[5050])
	loading_scheme = ['e','e','e','e','e','e']
	load = np.array([])
	#
	#t = np.append(t,np.append(t*2, np.append(t*3,t*4)))
	load = [ [ dt, dt, [g, 0.,0.,0.,0.,0.], loading_scheme ]  for g in GL_inp ]
	#
	print(len(load))
	return load
	
	
	# for testing elastoplastic model########################################################################################## loading_unloading #############################
def loading_unloading(time_factor = 1., total_stretch = 0.95, l_stretch_rate = -0.0005, ul_stretch_rate = 0.0005): # ramping_unloading
	#
	t_max_l = (total_stretch - 1.)/l_stretch_rate # total time for loading
	t_max_ul = (1.-total_stretch)/ul_stretch_rate # total time for unloading
	t_l = np.linspace(0,t_max_l,101)
	t_ul = np.linspace(0,t_max_ul,101)
	dt = t_max_l/len(t_l) # here I decided to only use loading time and disgard unloading time
	dt_ul =t_max_ul/len(t_ul) # then I changed my mind
	# lambda
	#
	GL_inp = np.array([])
	GL_11 = np.array([])
	delta_GL11 = np.array([0.])
	#dlambda = 0.
	stretch = 1.
	counter = 0
	#
	for i in range (len(t_l)):
		if (stretch > total_stretch):
			stretch = 1.+ l_stretch_rate*dt*i
			counter += 1
		CG = np.diag([stretch**2.,1./stretch,1./stretch])  ### different from Th PhD thesis, in this case for undrained compression.
		GL = 0.5*(CG - np.identity(3))
		GL_11 = np.append(GL_11,GL[0,0])
	print(stretch, len(t_l), len(GL_11))
	for j in range (len(t_ul)):
		if(stretch < 1.):
			stretch += 1.*ul_stretch_rate*dt_ul
		CG = np.diag([stretch**2.,1.,1.])
		GL = 0.5 *(CG - np.identity(3))
		GL_11 = np.append(GL_11,GL[0,0])
	print(stretch, len(t_ul),len(GL_11))
	#	
	for k in range(len(GL_11)-1):
		delta_GL11 = np.append(delta_GL11,GL_11[k+1]-GL_11[k])
	#
	loading_scheme = ['e','s','s','e','e','e']
	load = np.array([])
	#
	load = [ [ time_factor*dt, time_factor*dt, [g, 0.,0.,0.,0.,0.], loading_scheme ]  for g in delta_GL11 ]
	#
	#print(len(load))
	return load
	
#################################################################################### the end of loading_unloading ##################################################################
	
	
	
def delta_GL11_steps_unloading(dt,t_max,total_stretch,dlambdadt): # for single ramp and relaxation 
	#dt = (np.max(t)-np.min(t))/len(t)
	#dlambda = (total_stretch - 1.)/t_max
	nsteps=np.int(t_max/dt)



	stretch = 1. 
	delta_GL11 = np.array([0.])
	GL_11 = np.array([])
	counter = 1
	for i in range(nsteps):
		if (stretch > total_stretch):
			stretch = 1.+ dlambdadt*dt*i # dlambdadt is the stretch rate
			counter+=1
		CG = np.diag([stretch**2.,1.,1.])
		GL = 0.5*(CG - np.identity(3))
		GL_11 = np.append(GL_11, GL[0,0])
	print(counter)
	for j in range(len(GL_11)-1):
		delta_GL11 = np.append(delta_GL11,GL_11[j+1]-GL_11[j])
	return delta_GL11	
	
	

#stretch_torsion test
def haupt(time_factor=1.):
	#lambda
	l=np.array([])
	l = np.append(np.arange(1,1.8,0.08),l)
	l = np.append(l,np.arange(1.8,0.92,-0.08))
	#Dt
	d=np.array([])
	dmax = 50.*np.pi/180.
	d = np.append(np.arange(0,dmax,dmax/10),d)
	d = np.append(d,np.arange(dmax,-dmax/10,-dmax/10))
	#t
	t=np.array([])
	t = np.append(np.arange(0,10,1),t)
	t = np.append(t,np.arange(10,-1,-1))
	#
	r = 1./np.sqrt(l)
	loading_scheme = ['e','e','e','e','e','e'] #(shear: tensor)
	#load steps: time_interval, time_step_size, magnitude (11,22,33,12,23,13), type (..)
	load =  [[0.0, 1.0, np.array([0.0]*6), loading_scheme],
		[t[1]-t[0], t[1]-t[0], np.array([GL(calF(l,d,r,1))[0,0], 
			GL(calF(l,d,r,1))[1,1],
			GL(calF(l,d,r,1))[2,2],
			GL(calF(l,d,r,1))[0,1],
			GL(calF(l,d,r,1))[1,2],
			GL(calF(l,d,r,1))[0,2]]),loading_scheme],
		[t[2]-t[1], t[2]-t[1], np.array([GL(calF(l,d,r,2))[0,0]-GL(calF(l,d,r,1))[0,0], 
			GL(calF(l,d,r,2))[1,1]-GL(calF(l,d,r,1))[1,1],
			GL(calF(l,d,r,2))[2,2]-GL(calF(l,d,r,1))[2,2],
			GL(calF(l,d,r,2))[0,1]-GL(calF(l,d,r,1))[0,1],
			GL(calF(l,d,r,2))[1,2]-GL(calF(l,d,r,1))[1,2],
			GL(calF(l,d,r,2))[0,2]-GL(calF(l,d,r,1))[0,2]]),loading_scheme],
		[t[3]-t[2], t[3]-t[2], np.array([GL(calF(l,d,r,3))[0,0]-GL(calF(l,d,r,2))[0,0], 
			GL(calF(l,d,r,3))[1,1]-GL(calF(l,d,r,2))[1,1],
			GL(calF(l,d,r,3))[2,2]-GL(calF(l,d,r,2))[2,2],
			GL(calF(l,d,r,3))[0,1]-GL(calF(l,d,r,2))[0,1],
			GL(calF(l,d,r,3))[1,2]-GL(calF(l,d,r,2))[1,2],
			GL(calF(l,d,r,3))[0,2]-GL(calF(l,d,r,2))[0,2]]),loading_scheme],
		[t[4]-t[3], t[4]-t[3], np.array([GL(calF(l,d,r,4))[0,0]-GL(calF(l,d,r,3))[0,0], 
			GL(calF(l,d,r,4))[1,1]-GL(calF(l,d,r,3))[1,1],
			GL(calF(l,d,r,4))[2,2]-GL(calF(l,d,r,3))[2,2],
			GL(calF(l,d,r,4))[0,1]-GL(calF(l,d,r,3))[0,1],
			GL(calF(l,d,r,4))[1,2]-GL(calF(l,d,r,3))[1,2],
			GL(calF(l,d,r,4))[0,2]-GL(calF(l,d,r,3))[0,2]]),loading_scheme],
		[t[5]-t[4], t[5]-t[4], np.array([GL(calF(l,d,r,5))[0,0]-GL(calF(l,d,r,4))[0,0], 
			GL(calF(l,d,r,5))[1,1]-GL(calF(l,d,r,4))[1,1],
			GL(calF(l,d,r,5))[2,2]-GL(calF(l,d,r,4))[2,2],
			GL(calF(l,d,r,5))[0,1]-GL(calF(l,d,r,4))[0,1],
			GL(calF(l,d,r,5))[1,2]-GL(calF(l,d,r,4))[1,2],
			GL(calF(l,d,r,5))[0,2]-GL(calF(l,d,r,4))[0,2]]),loading_scheme],
		[t[6]-t[5], t[6]-t[5], np.array([GL(calF(l,d,r,6))[0,0]-GL(calF(l,d,r,5))[0,0], 
			GL(calF(l,d,r,6))[1,1]-GL(calF(l,d,r,5))[1,1],
			GL(calF(l,d,r,6))[2,2]-GL(calF(l,d,r,5))[2,2],
			GL(calF(l,d,r,6))[0,1]-GL(calF(l,d,r,5))[0,1],
			GL(calF(l,d,r,6))[1,2]-GL(calF(l,d,r,5))[1,2],
			GL(calF(l,d,r,6))[0,2]-GL(calF(l,d,r,5))[0,2]]),loading_scheme],
		[t[7]-t[6], t[7]-t[6], np.array([GL(calF(l,d,r,7))[0,0]-GL(calF(l,d,r,6))[0,0], 
			GL(calF(l,d,r,7))[1,1]-GL(calF(l,d,r,6))[1,1],
			GL(calF(l,d,r,7))[2,2]-GL(calF(l,d,r,6))[2,2],
			GL(calF(l,d,r,7))[0,1]-GL(calF(l,d,r,6))[0,1],
			GL(calF(l,d,r,7))[1,2]-GL(calF(l,d,r,6))[1,2],
			GL(calF(l,d,r,7))[0,2]-GL(calF(l,d,r,6))[0,2]]),loading_scheme],
		[t[8]-t[7], t[8]-t[7], np.array([GL(calF(l,d,r,8))[0,0]-GL(calF(l,d,r,7))[0,0], 
			GL(calF(l,d,r,8))[1,1]-GL(calF(l,d,r,7))[1,1],
			GL(calF(l,d,r,8))[2,2]-GL(calF(l,d,r,7))[2,2],
			GL(calF(l,d,r,8))[0,1]-GL(calF(l,d,r,7))[0,1],
			GL(calF(l,d,r,8))[1,2]-GL(calF(l,d,r,7))[1,2],
			GL(calF(l,d,r,8))[0,2]-GL(calF(l,d,r,7))[0,2]]),loading_scheme],
		[t[9]-t[8], t[9]-t[8], np.array([GL(calF(l,d,r,9))[0,0]-GL(calF(l,d,r,8))[0,0], 
			GL(calF(l,d,r,9))[1,1]-GL(calF(l,d,r,8))[1,1],
			GL(calF(l,d,r,9))[2,2]-GL(calF(l,d,r,8))[2,2],
			GL(calF(l,d,r,9))[0,1]-GL(calF(l,d,r,8))[0,1],
			GL(calF(l,d,r,9))[1,2]-GL(calF(l,d,r,8))[1,2],
			GL(calF(l,d,r,9))[0,2]-GL(calF(l,d,r,8))[0,2]]),loading_scheme],
		[t[10]-t[9], t[10]-t[9], np.array([GL(calF(l,d,r,10))[0,0]-GL(calF(l,d,r,9))[0,0], 
			GL(calF(l,d,r,10))[1,1]-GL(calF(l,d,r,9))[1,1],
			GL(calF(l,d,r,10))[2,2]-GL(calF(l,d,r,9))[2,2],
			GL(calF(l,d,r,10))[0,1]-GL(calF(l,d,r,9))[0,1],
			GL(calF(l,d,r,10))[1,2]-GL(calF(l,d,r,9))[1,2],
			GL(calF(l,d,r,10))[0,2]-GL(calF(l,d,r,9))[0,2]]),loading_scheme]]
	return load

def GL(F):
	return 0.5 * (np.dot(F.T,F) - np.eye(3))

def calF(l,d,r,i):
	F = np.eye(3)
	F[0,0] = 1./np.sqrt(l[i])
	F[1,1] = 1./np.sqrt(l[i])
	F[2,2] = l[i]
	F[1,2] = d[i] * r[i]
	return F




















